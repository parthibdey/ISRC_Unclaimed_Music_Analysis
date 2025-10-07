import os
import pandas as pd
import sqlite3
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import time
from dotenv import load_dotenv

# Load environment variables from .env
load_dotenv()


# STEP 1: Process Large TSV into SQLite


def create_isrc_database(tsv_path="C:/Users/parth/Downloads/unclaimedmusicalworkrightshares.tsv", 
                         db_path='unclaimed_works.db'):
    
    print("Creating SQLite database from TSV file")
    
    # Connect to SQLite
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Drop table if exists
    cursor.execute("DROP TABLE IF EXISTS unclaimed_works;")
    conn.commit()

    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS unclaimed_works (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            UnclaimedMusicalWorkRightShareRecordId TEXT,
            ResourceRecordId TEXT,
            MusicalWorkRecordId TEXT,
            ISRC TEXT,
            DspResourceId TEXT,
            ResourceTitle TEXT,
            ResourceSubTitle TEXT,
            AlternativeResourceTitle TEXT,
            DisplayArtistName TEXT,
            DisplayArtistISNI TEXT,
            Duration INTEGER,
            UnclaimedRightSharePercentage REAL,
            PercentileForPrioritisation REAL
        )
    ''')
    
    # Create index on ISRC for fast searching
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_isrc ON unclaimed_works(isrc)')
    
    # Process file in chunks
    chunk_size = 50000
    total_rows = 0
    
    for chunk in pd.read_csv("C:/Users/parth/Downloads/unclaimedmusicalworkrightshares.tsv", 
                             sep='\t', 
                             chunksize=chunk_size,
                             encoding='utf-8',
                             on_bad_lines='skip'):
        
        # Rename columns to match schema
        chunk.columns = [
            'UnclaimedMusicalWorkRightShareRecordId',
            'ResourceRecordId',
            'MusicalWorkRecordId',
            'ISRC',
            'DspResourceId',
            'ResourceTitle',
            'ResourceSubTitle',
            'AlternativeResourceTitle',
            'DisplayArtistName',
            'DisplayArtistISNI',
            'Duration',
            'UnclaimedRightSharePercentage',
            'PercentileForPrioritisation'
        ]
        
        # Insert into database
        chunk.to_sql('unclaimed_works', conn, if_exists='append', index=False)
        total_rows += len(chunk)
        print(f"Processed {total_rows:,} rows...")
    
    conn.commit()
    conn.close()
    print(f"\n✓ Database created successfully with {total_rows:,} records!")
    return total_rows



# STEP 2: Connect to Spotify API


def setup_spotify(client_id, client_secret):
    """Initialize Spotify API connection."""
    try:
        client_credentials_manager = SpotifyClientCredentials(
            client_id=client_id,
            client_secret=client_secret
        )
        sp = spotipy.Spotify(client_credentials_manager=client_credentials_manager)
        return sp
    except Exception as e:
        print(f"Error connecting to Spotify: {e}")
        
        return None


def get_artist_catalog(sp, artist_name):
    """
    Retrieve complete catalog for an artist from Spotify.
    Returns a DataFrame with all tracks and their ISRCs.
    """
    print(f"\nSearching for artist: {artist_name}")
    
    # Search for artist
    results = sp.search(q=f'artist:{artist_name}', type='artist', limit=1)
    
    if not results['artists']['items']:
        print(f"Artist '{artist_name}' not found!")
        return None
    
    artist = results['artists']['items'][0]
    artist_id = artist['id']
    artist_name_actual = artist['name']
    
    print(f"Found: {artist_name_actual}")
    print(f"Spotify URI: {artist['uri']}")
    print(f"Followers: {artist['followers']['total']:,}")
    print("\nRetrieving catalog:")
    
    catalog = []
    
    # Get all albums (includes albums, singles, compilations, appears_on)
    album_types = ['album', 'single', 'compilation']
    albums_seen = set()
    
    for album_type in album_types:
        offset = 0
        while True:
            try:
                results = sp.artist_albums(
                    artist_id, 
                    album_type=album_type, 
                    limit=50, 
                    offset=offset,
                    country='US'
                )
                
                for album in results['items']:
                    album_id = album['id']
                    
                    # Skip duplicates (same album in different markets)
                    if album_id in albums_seen:
                        continue
                    albums_seen.add(album_id)
                    
                    # Get all tracks from the album
                    album_details = sp.album(album_id)
                    
                    for track in album_details['tracks']['items']:
                        # Get full track details to access ISRC
                        track_full = sp.track(track['id'])
                        
                        isrc = track_full.get('external_ids', {}).get('isrc', 'N/A')
                        
                        catalog.append({
                            'track_name': track['name'],
                            'album': album['name'],
                            'album_type': album['album_type'],
                            'release_date': album['release_date'],
                            'isrc': isrc,
                            'duration_ms': track['duration_ms'],
                            'spotify_id': track['id'],
                            'spotify_url': track['external_urls']['spotify']
                        })
                        
                        print(f"  Added: {track['name']} [{isrc}]")
                
                if results['next'] is None:
                    break
                offset += 50
                time.sleep(0.1)  # Rate limiting
                
            except Exception as e:
                print(f"Error fetching albums: {e}")
                break
    
    df = pd.DataFrame(catalog)
    
    # Remove duplicates (same ISRC might appear multiple times)
    df = df.drop_duplicates(subset=['isrc'], keep='first')
    
    print(f"\n✓ Retrieved {len(df)} unique tracks")
    return df, artist_name_actual



# STEP 3: Cross-Reference with Database


def find_unclaimed_matches(artist_catalog_df, db_path):
    """
    Search the unclaimed works database for ISRCs from artist catalog.
    """
    print("\nSearching for matches in unclaimed works database...")
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    matches = []
    
    for idx, row in artist_catalog_df.iterrows():
        isrc = row['isrc']
        
        if isrc and isrc != 'N/A':
            # Search for ISRC in database
            cursor.execute('''
                SELECT * FROM unclaimed_works 
                WHERE isrc = ?
            ''', (isrc,))
            
            result = cursor.fetchone()
            
            if result:
                matches.append({
                    'track_name': row['track_name'],
                    'album': row['album'],
                    'release_date': row['release_date'],
                    'isrc': isrc,
                    'spotify_url': row['spotify_url'],
                    'unclaimed_title': result[6],  # resource_title
                    'unclaimed_artist': result[9],  # display_artist_name
                    'unclaimed_duration': result[11],  # duration
                    'unclaimed_record_id': result[1]  # record_id
                })
                print(f"  ✓ MATCH FOUND: {row['track_name']} [{isrc}]")
    
    conn.close()
    
    print(f"\n✓ Found {len(matches)} matches in unclaimed works!")
    return pd.DataFrame(matches) if matches else pd.DataFrame()



# STEP 4: Generate Excel Report


def create_excel_report(artist_catalog_df, matches_df, artist_name, output_file='music_rights_analysis.xlsx'):
    """
    Create a professional Excel report with multiple sheets.
    """
    print(f"\nGenerating Excel report: {output_file}")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Sheet 1: Artist Catalog
        artist_catalog_df.to_excel(writer, sheet_name='Artist Catalog', index=False)
        
        # Sheet 2: Unclaimed Matches
        if not matches_df.empty:
            matches_df.to_excel(writer, sheet_name='Unclaimed Matches', index=False)
        else:
            # Create empty sheet with message
            pd.DataFrame({
                'Message': ['No matches found in unclaimed works database']
            }).to_excel(writer, sheet_name='Unclaimed Matches', index=False)
        
        # Sheet 3: Analysis Summary
        summary_data = {
            'Metric': [
                'Artist Name',
                'Analysis Date',
                'Total Tracks in Catalog',
                'Tracks with ISRCs',
                'Tracks without ISRCs',
                'Matches Found in Unclaimed Works',
                'Match Rate',
                '',
                'Interpretation',
                '',
                'Database Info',
                'Spotify API',
                '',
                'Notes'
            ],
            'Value': [
                artist_name,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                len(artist_catalog_df),
                len(artist_catalog_df[artist_catalog_df['isrc'] != 'N/A']),
                len(artist_catalog_df[artist_catalog_df['isrc'] == 'N/A']),
                len(matches_df) if not matches_df.empty else 0,
                f"{(len(matches_df) / len(artist_catalog_df) * 100):.2f}%" if len(artist_catalog_df) > 0 and not matches_df.empty else "0%",
                '',
                'Matches indicate songs that may have unclaimed royalties or rights issues.',
                '',
                'Source: unclaimedmusicalworkrightshares.tsv (6.7GB)',
                'Retrieved via Spotipy library with full catalog access',
                '',
                'ISRC codes used as primary matching key. Analysis performed using SQLite for efficient searching.'
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Analysis Summary', index=False)
        
        # Format the sheets
        workbook = writer.book
        
        # Format Artist Catalog sheet
        ws1 = workbook['Artist Catalog']
        for cell in ws1[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Format Matches sheet
        if 'Unclaimed Matches' in workbook.sheetnames:
            ws2 = workbook['Unclaimed Matches']
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        
        # Format Summary sheet
        ws3 = workbook['Analysis Summary']
        for cell in ws3[1]:
            cell.font = Font(bold=True)
    
    print(f"✓ Excel report created: {output_file}")



# main 


def main():
    
    print("=" * 60)
    print("UNCLAIMED MUSIC RIGHTS ANALYSIS")
    print("=" * 60)
    
    
    TSV_FILE = os.getenv("TSV_FILE")
    DB_FILE = os.getenv("DB_FILE")
    output_dir = os.getenv("OUTPUT_DIR")
    
    
    
    ARTIST_NAME = 'Taylor Swift'  
    
    #comment out after running once
    #create_isrc_database(TSV_FILE, DB_FILE)
    
    
    # Connect to Spotify
    sp = setup_spotify(
        client_id=os.getenv("SPOTIFY_CLIENT_ID"),
        client_secret=os.getenv("SPOTIFY_CLIENT_SECRET")
    )
    if not sp:
        return
    
    # Get artist catalog
    result = get_artist_catalog(sp, ARTIST_NAME)
    if result is None:
        return
    
    artist_catalog_df, artist_name_actual = result
    
    # Find matches in unclaimed works
    matches_df = find_unclaimed_matches(artist_catalog_df, DB_FILE)
    
    # Generate Excel report
    output_filename = os.path.join(output_dir, f"{artist_name_actual.replace(' ', '_')}_rights_analysis(1).xlsx")
    create_excel_report(artist_catalog_df, matches_df, artist_name_actual, output_filename)
    
    print("\n" + "=" * 60)
    print("ANALYSIS COMPLETE!")
    print("=" * 60)
    print(f"Artist: {artist_name_actual}")
    print(f"Total tracks: {len(artist_catalog_df)}")
    print(f"Unclaimed matches: {len(matches_df)}")
    print(f"Output file: {output_filename}")
    print("=" * 60)


if __name__ == "__main__":
    main()